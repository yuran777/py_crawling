import requests
from bs4 import BeautifulSoup

#add code to use openpyxl
import openpyxl

try:
    wb = openpyxl.load_workbook("naver_article.xlsx")
    sheet = wb.active
    print("불러내기 완료")

except:
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["title", "press"])
    print("새로운 파일을 만들었습니다.")
    

#from requests import api
#wb = openpyxl.load_workbook("naver_article.xlsx")
#sheet = wb.active
#sheet.append(["title", "press"])
########

keyword = input("검색어를 입력하세요: ")

#f = open("naver_news2.csv", "w", encoding = "utf-8-sig" )
#f.write("title, source" + "\n")

for n in range(1, 100, 10):
    raw = requests.get("https://search.naver.com/search.naver?where=news&query="+keyword+"&start="+str(n))
    html = BeautifulSoup(raw.text, "html.parser")

    articles = html.select("ul.list_news> li")

    for art in articles:
        title = art.select_one("a.news_tit").text
        source = art.select_one("a.info.press").text

        #title = title.replace(",", "")
        #source = source.replace(",", "")
        print(title, source)

        sheet.append([title, source])

        #f.write(title + "," + source + "\n" )

#f.close()
wb.save("naver.xlsx")
